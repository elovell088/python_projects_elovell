#IT 414 - Eric Lovell - Excel Demo
import openpyxl

#Creating an instance of openpyxl
my_workbook = openpyxl.Workbook()

#Creating a workbook, naming it, and creating an additional sheet
curr_sheet = my_workbook.active
curr_sheet.title = "My Sheet"
my_workbook.create_sheet("Another Sheet")

#Inputtin a value into a specific sheet and cell
curr_sheet = my_workbook["Another Sheet"]
curr_sheet["B3"] = "My cell"

birthdays = [{"name" : "Chris", "birth_date" : "1/23/1975"},
             {"name" : "Diane", "birth_date" : "2/23/1982"},
             {"name" : "Jamal", "birth_date" : "4/21/1980"},
             {"name" : "Padma", "birth_date" : "11/07/1987"},]

curr_sheet = my_workbook["My Sheet"]
row_count = 1

for item in birthdays:
    col_count = 1
    curr_sheet.cell(row=row_count, column=col_count).value = item["name"]
    col_count += 1
    curr_sheet.cell(row=row_count, column=col_count).value = item["birth_date"]
    row_count += 1


curr_sheet.column_dimensions["B"].width = 50


my_workbook.save("files/my_workbook.xlsx")