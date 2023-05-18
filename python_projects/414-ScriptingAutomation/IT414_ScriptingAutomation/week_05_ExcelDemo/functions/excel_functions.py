#IT 414 - Eric Lovell - Excel Functions
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import *


def spreadSheetSetup(product_list, quantity_list):
    """"""   
    #Creating an instance of openpyxl
    my_workbook = openpyxl.Workbook()

    #Creating a workbook, naming it, and creating an additional sheet
    curr_sheet = my_workbook.active
    curr_sheet.title = "Sales Data Sheet"
    
    curr_sheet["A1"] = "My cell"

    products_and_quantity = []
    index_count = 0
    while index_count < len(product_list):
        temp_dict = {"product" : product_list[index_count], "quantity" : quantity_list[index_count]}
        products_and_quantity.append(temp_dict)
        index_count += 1
    
    curr_sheet.merge_cells("A1:B1")
    curr_sheet["A1"] = "Course Grading Data"

    curr_sheet["A1"].font = Font(b=True, color="33E8FF")
    curr_sheet["A1"].alignment = Alignment(horizontal="center")

    curr_sheet.cell(row=2, column=1).value = "Products"
    curr_sheet.cell(row=2, column=2).value = "Quantity"
    curr_sheet.cell(row=2, column=1).font = Font(b=True)
    curr_sheet.cell(row=2, column=2).font = Font(b=True)
    curr_sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    curr_sheet.cell(row=2, column=2).alignment = Alignment(horizontal="center")

    row_count = 3
    for item in products_and_quantity:
        col_count = 1
        curr_sheet.cell(row=row_count, column=col_count).value = item["product"]
        col_count += 1
        curr_sheet.cell(row=row_count, column=col_count).value = int(item["quantity"])
        row_count += 1


    curr_sheet.column_dimensions["B"].width = 20
    
    curr_sheet.freeze_panes = "A3"

    my_workbook.save("spreadsheets/sales_data.xlsx")

    

####################
def getAverage(quantity_list):
    """"""
    my_workbook = openpyxl.load_workbook("spreadsheets/sales_data.xlsx")


    curr_sheet = ["Sales Data Sheet"]
    max_row = curr_sheet.max_row
    max_col = curr_sheet.max_column

    row_count = 3
    
    curr_sheet["A" + str(max_row + 2)].value = "TOTAL"
    curr_sheet["B" + str(max_row + 2)].value = "=SUM(B" + str(row_count) + ":" + "B" + str(max_row) + ")"

    total_quantity = 0
    for item in quantity_list:
        total_quantity += int(item)

    row_count = 3
    items_found = []

    while row_count <= max_row:
        item = curr_sheet["A" + str(row_count)].value
        if item not in items_found:
            items_found.append(item)

        row_count +=1 
    
    last_col_letter = get_column_letter(max_col)
    next_col_letter = get_column_letter(max_col + 1)
    another_col_letter = get_column_letter(max_col + 2)

    col_count_new = 0
    row_count = 3
    new_total = 0
    for item in items_found:
        total = 0
        col_count_new += 1
        while row_count <= max_row:
            product_item = curr_sheet["A" + str(row_count)].value
            if item == product_item:
                quantity = int(curr_sheet["B" + str(row_count)].value)
                total += quantity
                print(total)
            row_count += 1
            
            curr_sheet[next_col_letter]
            curr_sheet[next_col_letter + str(max_row + 2)].value = item + str(total/total_quantity)
    
    my_workbook.save("files/gradebook_reformat.xlsx")